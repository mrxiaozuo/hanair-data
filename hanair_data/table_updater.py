"""Download the HNAir table and update an Excel workbook.

The module exposes two main helpers:

```
fetch_table_rows(url: str, table_index: int = 0) -> list[list[str]]
update_workbook(rows: Sequence[Sequence[str]], output_path: PathLike, ...)
```

A small command line interface is also provided, so the script can be executed
with ``python -m hanair_data.table_updater``.
"""

from __future__ import annotations

import argparse
import sys
from dataclasses import dataclass
from datetime import datetime
from html.parser import HTMLParser
from pathlib import Path
from typing import List, Sequence
from urllib.error import HTTPError, URLError
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from . import DEFAULT_TABLE_URL

USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36"
)


class _TableParser(HTMLParser):
    """Extract a table from HTML using the standard library."""

    def __init__(self, target_index: int = 0):
        super().__init__()
        if target_index < 0:
            raise ValueError("table index must be >= 0")
        self._target_index = target_index
        self._current_index = -1
        self._capture = False
        self._table_depth = 0
        self._in_row = False
        self._in_cell = False
        self._rows: List[List[str]] = []
        self._current_row: List[str] = []
        self._cell_parts: List[str] = []

    def handle_starttag(self, tag: str, attrs):  # type: ignore[override]
        if tag.lower() == "table":
            self._current_index += 1
            if self._current_index == self._target_index:
                self._capture = True
                self._table_depth = 1
                return
            if self._capture:
                self._table_depth += 1
                return

        if not self._capture:
            return

        if tag.lower() == "tr":
            self._in_row = True
            self._current_row = []
        elif tag.lower() in {"td", "th"}:
            self._in_cell = True
            self._cell_parts = []
        elif tag.lower() == "br" and self._in_cell:
            self._cell_parts.append("\n")

    def handle_endtag(self, tag: str):  # type: ignore[override]
        tag = tag.lower()
        if tag == "table" and self._capture:
            self._table_depth -= 1
            if self._table_depth == 0:
                self._capture = False
            return

        if not self._capture:
            return

        if tag in {"td", "th"} and self._in_cell:
            text = "".join(self._cell_parts).replace("\xa0", " ")
            text = "\n".join(part.strip() for part in text.splitlines())
            text = " ".join(text.split()) if "\n" not in text else text
            self._current_row.append(text.strip())
            self._in_cell = False
            self._cell_parts = []
        elif tag == "tr" and self._in_row:
            if self._current_row:
                self._rows.append(self._current_row)
            self._in_row = False
            self._current_row = []

    def handle_data(self, data: str):  # type: ignore[override]
        if self._capture and self._in_cell:
            self._cell_parts.append(data)

    def handle_startendtag(self, tag: str, attrs):  # type: ignore[override]
        if tag.lower() == "br" and self._capture and self._in_cell:
            self._cell_parts.append("\n")

    @property
    def rows(self) -> List[List[str]]:
        return self._rows


@dataclass(frozen=True)
class FetchResult:
    rows: List[List[str]]
    fetched_at: datetime
    url: str


def fetch_table_rows(url: str, table_index: int = 0, *, timeout: float = 30.0) -> FetchResult:
    """Download a table and return its rows.

    Parameters
    ----------
    url:
        Page containing the target table.
    table_index:
        Zero-based index of the table to extract.
    timeout:
        Timeout (in seconds) for the network request.
    """

    request = Request(url, headers={"User-Agent": USER_AGENT})
    with urlopen(request, timeout=timeout) as response:
        encoding = response.headers.get_content_charset() or "utf-8"
        html = response.read().decode(encoding, errors="ignore")

    parser = _TableParser(table_index=table_index)
    parser.feed(html)
    rows = parser.rows
    if not rows:
        raise ValueError(
            "No table rows were found. Check that the page structure has not changed."
        )

    return FetchResult(rows=rows, fetched_at=datetime.now(), url=url)


def _load_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    return Workbook()


def _reset_sheet(workbook: Workbook, name: str):
    if len(name) > 31:
        raise ValueError("Excel sheet names must be 31 characters or fewer")

    if (
        workbook.sheetnames == ["Sheet"]
        and workbook["Sheet"].max_row == 1
        and workbook["Sheet"].max_column == 1
        and workbook["Sheet"].cell(1, 1).value in (None, "")
    ):
        worksheet = workbook.active
        worksheet.title = name
        return worksheet

    if name in workbook.sheetnames:
        worksheet = workbook[name]
        workbook.remove(worksheet)

    return workbook.create_sheet(title=name)


def _write_rows(worksheet, rows: Sequence[Sequence[str]]):
    column_widths: dict[int, int] = {}
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            if value is None:
                text = ""
            else:
                text = str(value)
            worksheet.cell(row=row_idx, column=col_idx, value=text)
            width = max(len(segment) for segment in text.splitlines()) if text else 0
            column_widths[col_idx] = max(column_widths.get(col_idx, 0), width)

    for col_idx, width in column_widths.items():
        adjusted = min(max(width + 2, 10), 60)
        worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted

    worksheet.freeze_panes = "A2"


def update_workbook(
    rows: Sequence[Sequence[str]],
    output_path: Path | str,
    *,
    latest_sheet_name: str = "Latest",
    include_history: bool = True,
    history_sheet_name: str | None = None,
    fetched_at: datetime | None = None,
    source_url: str | None = None,
) -> Path:
    """Write rows into an Excel workbook.

    Parameters
    ----------
    rows:
        Table data to write.
    output_path:
        File that should contain the workbook.
    latest_sheet_name:
        Sheet that always reflects the most recent download.
    include_history:
        When ``True`` a dated sheet is added on each run.
    history_sheet_name:
        Optional explicit sheet name for the historical copy. If omitted the
        current date (``YYYY-MM-DD``) is used.
    fetched_at:
        Timestamp used when annotating the workbook metadata.
    """

    if not rows:
        raise ValueError("No rows provided for Excel export")

    timestamp = fetched_at or datetime.now()
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    workbook = _load_workbook(target)

    latest_sheet = _reset_sheet(workbook, latest_sheet_name)
    _write_rows(latest_sheet, rows)

    if include_history:
        history_name = history_sheet_name or timestamp.strftime("%Y-%m-%d")
        if len(history_name) > 31:
            raise ValueError("History sheet name is longer than Excel's 31 character limit")
        history_sheet = _reset_sheet(workbook, history_name)
        _write_rows(history_sheet, rows)

    if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
        sheet = workbook["Sheet"]
        if not any(cell.value for row in sheet.iter_rows() for cell in row):
            workbook.remove(sheet)

    workbook.properties.modified = timestamp
    workbook.properties.lastModifiedBy = "hanair-data automation"
    workbook.properties.description = (
        f"Data fetched from {(source_url or DEFAULT_TABLE_URL)} "
        f"on {timestamp.isoformat(timespec='seconds')}"
    )

    workbook.save(target)
    return target


def build_argument_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Download the HNAir table and update an Excel workbook. "
            "If the workbook already exists, the 'Latest' sheet is replaced and a "
            "dated history sheet is added (unless disabled)."
        )
    )
    parser.add_argument(
        "--url",
        default=DEFAULT_TABLE_URL,
        help="Page containing the target table (defaults to the known announcement URL).",
    )
    parser.add_argument(
        "--output",
        default="hnair_table.xlsx",
        help="Excel workbook to update (defaults to ./hnair_table.xlsx)",
    )
    parser.add_argument(
        "--table-index",
        type=int,
        default=0,
        help="Zero-based index of the table to extract when multiple tables are present.",
    )
    parser.add_argument(
        "--latest-sheet-name",
        default="Latest",
        help="Name of the sheet that stores the most recent snapshot.",
    )
    parser.add_argument(
        "--history-sheet-name",
        default=None,
        help=(
            "Optional explicit name for the historical sheet. By default the sheet is "
            "named with today's date (YYYY-MM-DD)."
        ),
    )
    parser.add_argument(
        "--skip-history",
        action="store_true",
        help="Only update the latest sheet without keeping dated history sheets.",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=30.0,
        help="Timeout (in seconds) for the HTTP request.",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_argument_parser()
    args = parser.parse_args(argv)

    try:
        fetch_result = fetch_table_rows(args.url, table_index=args.table_index, timeout=args.timeout)
    except (HTTPError, URLError) as exc:
        parser.error(f"Failed to download table: {exc}")
        return 2
    except ValueError as exc:
        parser.error(str(exc))
        return 3

    try:
        output_path = update_workbook(
            fetch_result.rows,
            args.output,
            latest_sheet_name=args.latest_sheet_name,
            include_history=not args.skip_history,
            history_sheet_name=args.history_sheet_name,
            fetched_at=fetch_result.fetched_at,
            source_url=fetch_result.url,
        )
    except ValueError as exc:
        parser.error(str(exc))
        return 4

    print(
        f"Saved {len(fetch_result.rows)} rows to '{output_path}'. "
        f"Fetched at {fetch_result.fetched_at.isoformat(timespec='seconds')} from {fetch_result.url}."
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
